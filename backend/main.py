import os
import re
import io
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill
from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware

app = FastAPI()

# Permitir peticiones desde tu frontend
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"], # O la URL de tu frontend en Vercel
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ----------------------------------------------------
# FUNCIONES DE MANTENIMIENTO DE LÓGICA (Tus funciones)
# ----------------------------------------------------

def normalizar_texto(texto):
    if not texto:
        return ""
    texto = str(texto).strip().lower()
    reemplazos = {
        'á': 'a', 'é': 'e', 'í': 'i', 'ó': 'o', 'ú': 'u',
        'ñ': 'n', '_': ' ', '-': ' ', 'º': '', '°': '', '#': '', '.': ''
    }
    for orig, repl in reemplazos.items():
        texto = texto.replace(orig, repl)
    return re.sub(r'\s+', ' ', texto).strip()

def buscar_columna(headers, nombres_posibles):
    for idx, header in enumerate(headers):
        if header:
            header_normalizado = normalizar_texto(header)
            for posible in nombres_posibles:
                posible_normalizado = normalizar_texto(posible)
                if posible_normalizado == header_normalizado or posible_normalizado in header_normalizado:
                    return idx + 1
    return None

def es_numerico(valor):
    if valor is None:
        return False
    try:
        float(valor)
        return True
    except ValueError:
        return False

def procesar_y_limpiar_fecha(val_fecha):
    if val_fecha is None:
        return None, False
    if isinstance(val_fecha, datetime):
        return val_fecha.strftime("%Y%m%d"), True
    str_fecha = str(val_fecha).strip()
    if str_fecha.endswith('.0'):
        str_fecha = str_fecha[:-2]
    if not re.match(r'^\d{8}$', str_fecha):
        return str_fecha, False
    try:
        datetime.strptime(str_fecha, "%Y%m%d")
        return str_fecha, True
    except ValueError:
        return str_fecha, False

def es_fila_vacia(valores):
    return all(v is None or str(v).strip() == "" for v in valores)

def encontrar_mejor_hoja_y_cabecera(wb_orig):
    mejor_ws = None
    mejor_fila_cabecera = 1
    mejor_mapeo = {}
    max_coincidencias = 0

    columnas_a_buscar = {
        "tipo_id": ["tipo de identificacion", "tipo de documento", "tipo id", "tipo_identificacion"],
        "n_id": ["n identificacion", "nº identificacion", "no identificacion", "numero identificacion"],
        "nombre": ["nombre tercero", "tercero"],
        "obligacion": ["numero obligacion", "obligacion"],
        "edad_mora": ["edad de mora", "edad mora"],
        "cuotas_mora": ["cuotas de mora", "cuotas en mora", "cuotas mora"],
        "val_inicial": ["valor inicial"],
        "val_mora": ["valor de mora", "valor en mora"],
        "fecha_inicio": ["fecha inicio"],
        "fecha_corte": ["fecha de corte", "fecha corte"]
    }

    for sheet_name in wb_orig.sheetnames:
        ws = wb_orig[sheet_name]
        limite_filas = min(15, ws.max_row + 1)
        for r in range(1, limite_filas):
            headers = [cell.value for cell in ws[r]]
            mapeo_actual = {}
            coincidencias = 0
            for clave, terminos in columnas_a_buscar.items():
                idx = buscar_columna(headers, terminos)
                mapeo_actual[clave] = idx
                if idx is not None:
                    coincidencias += 1

            if coincidencias > max_coincidencias:
                max_coincidencias = coincidencias
                mejor_ws = ws
                mejor_fila_cabecera = r
                mejor_mapeo = mapeo_actual

    return mejor_ws, mejor_fila_cabecera, mejor_mapeo, max_coincidencias

# ----------------------------------------------------
# ENDPOINT DE FASTAPI PARA RECIBIR Y DEVOLVER EL EXCEL
# ----------------------------------------------------

@app.post("/procesar-excel/")
async def procesar_excel_endpoint(file: UploadFile = File(...)):
    contents = await file.read()
    wb_orig = openpyxl.load_workbook(filename=io.BytesIO(contents), data_only=True)

    ws_orig, fila_cabecera, mapeo, coincidencias = encontrar_mejor_hoja_y_cabecera(wb_orig)

    if not ws_orig or coincidencias < 4:
        raise HTTPException(status_code=400, detail="No se encontró una hoja que contenga las columnas requeridas.")

    wb_nuevo = openpyxl.Workbook()
    ws_nuevo = wb_nuevo.active
    ws_nuevo.title = "Revisión Simplificada"

    nuevas_cabeceras = [
        "Tipo de identificación", "Nº identificación", "Nombre tercero", "Numero obligacion",
        "edad de mora", "cuotas en mora", "Valor inicial",
        "Valor de mora", "Fecha inicio", "Fecha de corte",
        "Inconsistencia Detectada"
    ]
    ws_nuevo.append(nuevas_cabeceras)

    fill_rojo = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
    fill_naranja = PatternFill(start_color="FFFFEB9C", end_color="FFFFEB9C", fill_type="solid")
    fill_amarillo = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")

    idx_tipo_id = mapeo.get("tipo_id")
    idx_id = mapeo.get("n_id")
    idx_nombre = mapeo.get("nombre")
    idx_obli = mapeo.get("obligacion")
    idx_edad = mapeo.get("edad_mora")
    idx_cuotas = mapeo.get("cuotas_mora")
    idx_val_ini = mapeo.get("val_inicial")
    idx_val_mor = mapeo.get("val_mora")
    idx_f_ini = mapeo.get("fecha_inicio")
    idx_f_cor = mapeo.get("fecha_corte")

    fila_salida = 2
    for row in range(fila_cabecera + 1, ws_orig.max_row + 1):
        val_tipo_id = ws_orig.cell(row=row, column=idx_tipo_id).value if idx_tipo_id else None
        val_id = ws_orig.cell(row=row, column=idx_id).value if idx_id else None
        val_nombre = ws_orig.cell(row=row, column=idx_nombre).value if idx_nombre else None
        val_obligacion = ws_orig.cell(row=row, column=idx_obli).value if idx_obli else None
        val_edad = ws_orig.cell(row=row, column=idx_edad).value if idx_edad else None
        val_cuotas = ws_orig.cell(row=row, column=idx_cuotas).value if idx_cuotas else None
        val_inicial = ws_orig.cell(row=row, column=idx_val_ini).value if idx_val_ini else None
        val_mora = ws_orig.cell(row=row, column=idx_val_mor).value if idx_val_mor else None
        val_f_inicio = ws_orig.cell(row=row, column=idx_f_ini).value if idx_f_ini else None
        val_f_corte = ws_orig.cell(row=row, column=idx_f_cor).value if idx_f_cor else None

        if es_fila_vacia([val_tipo_id, val_id, val_nombre, val_obligacion, val_edad, val_cuotas, val_inicial, val_mora, val_f_inicio, val_f_corte]):
            continue

        f_inicio_texto, inicio_ok = procesar_y_limpiar_fecha(val_f_inicio)
        f_corte_texto, corte_ok = procesar_y_limpiar_fecha(val_f_corte)

        ws_nuevo.cell(row=fila_salida, column=1, value=val_tipo_id)
        ws_nuevo.cell(row=fila_salida, column=2, value=val_id)
        ws_nuevo.cell(row=fila_salida, column=3, value=val_nombre)
        ws_nuevo.cell(row=fila_salida, column=4, value=val_obligacion)
        ws_nuevo.cell(row=fila_salida, column=5, value=val_edad)
        ws_nuevo.cell(row=fila_salida, column=6, value=val_cuotas)
        ws_nuevo.cell(row=fila_salida, column=7, value=val_inicial)
        ws_nuevo.cell(row=fila_salida, column=8, value=val_mora)
        ws_nuevo.cell(row=fila_salida, column=9, value=f_inicio_texto)
        ws_nuevo.cell(row=fila_salida, column=10, value=f_corte_texto)

        errores_fila = []
        tiene_error_rojo = False
        tiene_error_naranja = False
        tiene_error_amarillo = False

        if idx_tipo_id:
            if val_tipo_id is not None:
                str_tipo_id = str(val_tipo_id).strip()
                if str_tipo_id in ["00", "0", "0.0"]:
                    ws_nuevo.cell(row=fila_salida, column=1).fill = fill_amarillo
                    errores_fila.append("responsabilidad asesora {tipo de identificación mala}")
                    tiene_error_amarillo = True
            else:
                ws_nuevo.cell(row=fila_salida, column=1).fill = fill_naranja
                errores_fila.append("Tipo de Identificación vacío")
                tiene_error_naranja = True

        if idx_id:
            if val_id is not None:
                str_id = str(val_id).strip()
                if str_id in ["00", "0", "0.0"]:
                    ws_nuevo.cell(row=fila_salida, column=2).fill = fill_rojo
                    errores_fila.append("Nº Identificación es '00' o '0' (Asesora)")
                    tiene_error_rojo = True
            else:
                ws_nuevo.cell(row=fila_salida, column=2).fill = fill_naranja
                errores_fila.append("Nº Identificación vacío")
                tiene_error_naranja = True

        if idx_f_ini and not inicio_ok:
            ws_nuevo.cell(row=fila_salida, column=9).fill = fill_naranja
            errores_fila.append("Formato incorrecto en Fecha Inicio (debe ser AAAAMMDD)")
            tiene_error_naranja = True

        if idx_f_cor and not corte_ok:
            ws_nuevo.cell(row=fila_salida, column=10).fill = fill_naranja
            errores_fila.append("Formato incorrecto en Fecha Corte (debe ser AAAAMMDD)")
            tiene_error_naranja = True

        if idx_f_ini and idx_f_cor and inicio_ok and corte_ok:
            try:
                dt_inicio = datetime.strptime(f_inicio_texto, "%Y%m%d")
                dt_corte = datetime.strptime(f_corte_texto, "%Y%m%d")
                if dt_inicio > dt_corte:
                    ws_nuevo.cell(row=fila_salida, column=9).fill = fill_naranja
                    ws_nuevo.cell(row=fila_salida, column=10).fill = fill_naranja
                    errores_fila.append("Fecha inicio es posterior a Fecha de corte")
                    tiene_error_naranja = True
            except Exception:
                pass

        inicial_ok = es_numerico(val_inicial) if idx_val_ini else False
        mora_ok = es_numerico(val_mora) if idx_val_mor else False

        if idx_val_ini and not inicial_ok:
            ws_nuevo.cell(row=fila_salida, column=7).fill = fill_naranja
            errores_fila.append("Valor inicial vacío o no es numérico")
            tiene_error_naranja = True

        if idx_val_mor and not mora_ok:
            ws_nuevo.cell(row=fila_salida, column=8).fill = fill_naranja
            errores_fila.append("Valor de mora vacío o no es numérico")
            tiene_error_naranja = True

        if idx_val_ini and idx_val_mor and inicial_ok and mora_ok:
            if float(val_inicial) < float(val_mora):
                ws_nuevo.cell(row=fila_salida, column=7).fill = fill_naranja
                ws_nuevo.cell(row=fila_salida, column=8).fill = fill_naranja
                errores_fila.append("Valor inicial es menor al Valor de mora")
                tiene_error_naranja = True

        if idx_edad and idx_cuotas:
            difieren = False
            if val_edad is not None and val_cuotas is not None:
                if es_numerico(val_edad) and es_numerico(val_cuotas):
                    if float(val_edad) != float(val_cuotas):
                        difieren = True
                        errores_fila.append(f"Edad de mora ({int(float(val_edad))}) no coincide con Cuotas en mora ({int(float(val_cuotas))})")
                else:
                    difieren = True
                    errores_fila.append("Edad o Cuotas de mora no son numéricos")
            elif val_edad is not None or val_cuotas is not None:
                difieren = True
                errores_fila.append("Uno de los campos de mora (Edad o Cuotas) está vacío")

            if difieren:
                ws_nuevo.cell(row=fila_salida, column=5).fill = fill_rojo
                ws_nuevo.cell(row=fila_salida, column=6).fill = fill_rojo
                tiene_error_rojo = True

        if errores_fila:
            ws_nuevo.cell(row=fila_salida, column=11, value=" | ".join(errores_fila))
            if tiene_error_rojo:
                ws_nuevo.cell(row=fila_salida, column=11).fill = fill_rojo
            elif tiene_error_amarillo:
                ws_nuevo.cell(row=fila_salida, column=11).fill = fill_amarillo
            elif tiene_error_naranja:
                ws_nuevo.cell(row=fila_salida, column=11).fill = fill_naranja
        else:
            ws_nuevo.cell(row=fila_salida, column=11, value="Sin novedades")

        fila_salida += 1

    # Guardar en memoria y retornar el archivo descargable
    output = io.BytesIO()
    wb_nuevo.save(output)
    output.seek(0)

    headers = {'Content-Disposition': 'attachment; filename="Resultado_Revision_Simplificado.xlsx"'}
    return StreamingResponse(output, headers=headers, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")