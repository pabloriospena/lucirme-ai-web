import os
import uuid
import shutil
import pandas as pd
# pyrefly: ignore [missing-import]
from fastapi import APIRouter, File, UploadFile, HTTPException
# pyrefly: ignore [missing-import]
from fastapi.responses import FileResponse
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, numbers

router = APIRouter()

# Colores para la conciliación
FILL_VERDE = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_AMARILLO = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_NARANJA = PatternFill(start_color="FCD5B4", end_color="FCD5B4", fill_type="solid")
FILL_ROJO = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_AZUL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
FILL_HEADER = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
FONT_HEADER = Font(color="FFFFFF", bold=True, size=11)

UMBRAL_COINCIDE = 1000
UMBRAL_DIFERENCIA_MENOR = 100000


def limpiar_numero(valor):
    """Convierte valores a float, manejando strings con comas o símbolos."""
    if pd.isna(valor) or valor is None:
        return 0.0
    if isinstance(valor, (int, float)):
        return float(valor)
    try:
        return float(str(valor).replace(",", "").replace("$", "").strip())
    except (ValueError, TypeError):
        return 0.0


def limpiar_nit(valor):
    """Limpia el NIT quitando puntos, guiones y espacios."""
    if pd.isna(valor) or valor is None:
        return ""
    return str(valor).replace(".", "").replace("-", "").replace(" ", "").strip()


def leer_facturas_dian(ruta_archivo: str) -> pd.DataFrame:
    """Lee y limpia el archivo de facturas DIAN."""
    df = pd.read_excel(ruta_archivo, engine="openpyxl")
    
    # Filtrar solo documentos relevantes (excluir "Application response")
    tipos_validos = [
        "Factura electrónica",
        "Nota de crédito electrónica",
        "Documento equivalente POS",
        "Factura electrónica de contingencia"
    ]
    df = df[df["Tipo de documento"].isin(tipos_validos)].copy()
    
    # Limpiar NIT Emisor
    df["NIT Emisor Limpio"] = df["NIT Emisor"].apply(limpiar_nit)
    df["Total Limpio"] = df["Total"].apply(limpiar_numero)
    
    return df


def leer_movimientos_contables(ruta_archivo: str) -> pd.DataFrame:
    """Lee y limpia el archivo de movimientos contables."""
    df = pd.read_excel(ruta_archivo, engine="openpyxl")
    
    # Limpiar Identificación y valores
    df["Identificacion Limpia"] = df["Identificación"].apply(limpiar_nit)
    df["Debito Limpio"] = df["Débito"].apply(limpiar_numero)
    df["Credito Limpio"] = df["Crédito"].apply(limpiar_numero)
    
    return df


def conciliar_por_tercero(df_facturas: pd.DataFrame, df_movimientos: pd.DataFrame) -> pd.DataFrame:
    """Agrupa y concilia por tercero."""
    # Agrupar facturas por NIT Emisor
    facturas_agg = df_facturas.groupby("NIT Emisor Limpio").agg(
        Nombre_Tercero=("Nombre Emisor", "first"),
        Total_Facturas=("Total Limpio", "sum")
    ).reset_index()
    facturas_agg.rename(columns={"NIT Emisor Limpio": "NIT"}, inplace=True)
    
    # Agrupar movimientos por Identificación
    movimientos_agg = df_movimientos.groupby("Identificacion Limpia").agg(
        Nombre_Tercero_Cont=("Nombre tercero", "first"),
        Total_Debito=("Debito Limpio", "sum"),
        Total_Credito=("Credito Limpio", "sum")
    ).reset_index()
    movimientos_agg.rename(columns={"Identificacion Limpia": "NIT"}, inplace=True)
    movimientos_agg["Neto_Contable"] = movimientos_agg["Total_Debito"] - movimientos_agg["Total_Credito"]
    
    # Full outer join
    conciliacion = pd.merge(facturas_agg, movimientos_agg, on="NIT", how="outer", suffixes=("_Fac", "_Cont"))
    
    # Unificar nombre: preferir el de facturación, si no existe usar el contable
    conciliacion["Nombre"] = conciliacion["Nombre_Tercero_Fac"].fillna(conciliacion["Nombre_Tercero_Cont"])
    
    # Llenar NaN con 0
    conciliacion["Total_Facturas"] = conciliacion["Total_Facturas"].fillna(0)
    conciliacion["Total_Debito"] = conciliacion["Total_Debito"].fillna(0)
    conciliacion["Total_Credito"] = conciliacion["Total_Credito"].fillna(0)
    conciliacion["Neto_Contable"] = conciliacion["Neto_Contable"].fillna(0)
    
    # Calcular diferencia absoluta
    conciliacion["Diferencia"] = (conciliacion["Total_Facturas"] - conciliacion["Neto_Contable"]).abs()
    
    # Clasificar estado
    def clasificar(row):
        if row["Total_Facturas"] == 0 and row["Neto_Contable"] != 0:
            return "Solo en Contabilidad"
        if row["Total_Facturas"] > 0 and row["Neto_Contable"] == 0:
            return "Solo en Facturación"
        if row["Diferencia"] < UMBRAL_COINCIDE:
            return "Coincide"
        if row["Diferencia"] < UMBRAL_DIFERENCIA_MENOR:
            return "Diferencia Menor"
        return "Diferencia"
    
    conciliacion["Estado"] = conciliacion.apply(clasificar, axis=1)
    
    # Ordenar: primero los que tienen problemas, luego los que coinciden
    orden_estado = {"Diferencia": 0, "Solo en Facturación": 1, "Solo en Contabilidad": 2, 
                    "Diferencia Menor": 3, "Coincide": 4}
    conciliacion["orden"] = conciliacion["Estado"].map(orden_estado)
    conciliacion = conciliacion.sort_values(["orden", "Diferencia"], ascending=[True, False])
    
    # Seleccionar columnas finales
    resultado = conciliacion[[
        "NIT", "Nombre", "Total_Facturas", "Total_Debito", "Total_Credito", 
        "Neto_Contable", "Diferencia", "Estado"
    ]].copy()
    
    return resultado


def escribir_excel_conciliacion(ruta_salida: str, df_conciliacion: pd.DataFrame, 
                                 df_facturas: pd.DataFrame, df_movimientos: pd.DataFrame):
    """Escribe el Excel final con 3 hojas y colores."""
    wb = Workbook()
    
    # ===== HOJA 1: CONCILIACIÓN POR TERCERO =====
    ws1 = wb.active
    ws1.title = "Conciliación por Tercero"
    
    columnas_conciliacion = [
        "NIT", "Nombre Tercero", "Total Facturas DIAN", 
        "Total Débito Contable", "Total Crédito Contable", 
        "Neto Contable (Déb - Créd)", "Diferencia", "Estado"
    ]
    ws1.append(columnas_conciliacion)
    
    # Formato del header
    for col_idx, _ in enumerate(columnas_conciliacion, start=1):
        cell = ws1.cell(row=1, column=col_idx)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Mapeo de estado a color
    estado_color = {
        "Coincide": FILL_VERDE,
        "Diferencia Menor": FILL_AMARILLO,
        "Diferencia": FILL_NARANJA,
        "Solo en Facturación": FILL_ROJO,
        "Solo en Contabilidad": FILL_AZUL
    }
    
    # Escribir datos y pintar
    for row_idx, row in df_conciliacion.iterrows():
        fila = [
            row["NIT"],
            row["Nombre"],
            row["Total_Facturas"],
            row["Total_Debito"],
            row["Total_Credito"],
            row["Neto_Contable"],
            row["Diferencia"],
            row["Estado"]
        ]
        ws1.append(fila)
        excel_row = ws1.max_row
        
        color = estado_color.get(row["Estado"], FILL_VERDE)
        for col_idx in range(1, len(columnas_conciliacion) + 1):
            cell = ws1.cell(row=excel_row, column=col_idx)
            cell.fill = color
            # Formato numérico para columnas de dinero (3 a 7)
            if 3 <= col_idx <= 7:
                cell.number_format = '#,##0.00'
    
    # Ajustar anchos
    ws1.column_dimensions["A"].width = 15  # NIT
    ws1.column_dimensions["B"].width = 40  # Nombre
    for col in ["C", "D", "E", "F", "G"]:
        ws1.column_dimensions[col].width = 20
    ws1.column_dimensions["H"].width = 25  # Estado
    
    # ===== HOJA 2: DETALLE FACTURAS DIAN =====
    ws2 = wb.create_sheet("Detalle facturas DIAN")
    columnas_facturas = [
        "Tipo de documento", "Prefijo", "Folio", "Fecha Emisión",
        "NIT Emisor", "Nombre Emisor", "NIT Receptor", "Nombre Receptor", "Total", "Estado"
    ]
    ws2.append(columnas_facturas)
    for col_idx, _ in enumerate(columnas_facturas, start=1):
        cell = ws2.cell(row=1, column=col_idx)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
    
    for _, row in df_facturas.iterrows():
        fila = [
            row.get("Tipo de documento", ""),
            row.get("Prefijo", ""),
            row.get("Folio", ""),
            row.get("Fecha Emisión", ""),
            row.get("NIT Emisor", ""),
            row.get("Nombre Emisor", ""),
            row.get("NIT Receptor", ""),
            row.get("Nombre Receptor", ""),
            row.get("Total", 0),
            row.get("Estado", "")
        ]
        ws2.append(fila)
        excel_row = ws2.max_row
        # Formato numérico en Total
        ws2.cell(row=excel_row, column=9).number_format = '#,##0.00'
    
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 10
    ws2.column_dimensions["C"].width = 12
    ws2.column_dimensions["D"].width = 15
    ws2.column_dimensions["E"].width = 15
    ws2.column_dimensions["F"].width = 35
    ws2.column_dimensions["G"].width = 15
    ws2.column_dimensions["H"].width = 25
    ws2.column_dimensions["I"].width = 18
    ws2.column_dimensions["J"].width = 25
    
    # ===== HOJA 3: DETALLE MOVIMIENTOS =====
    ws3 = wb.create_sheet("Detalle movimientos")
    columnas_movimientos = [
        "Cuenta contable", "Identificación", "Nombre tercero",
        "Comprobante", "Fecha elaboración", "Descripción", "Débito", "Crédito"
    ]
    ws3.append(columnas_movimientos)
    for col_idx, _ in enumerate(columnas_movimientos, start=1):
        cell = ws3.cell(row=1, column=col_idx)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
    
    for _, row in df_movimientos.iterrows():
        fila = [
            row.get("Cuenta contable", ""),
            row.get("Identificación", ""),
            row.get("Nombre tercero", ""),
            row.get("Comprobante", ""),
            row.get("Fecha elaboración", ""),
            row.get("Descripción", ""),
            row.get("Débito", 0),
            row.get("Crédito", 0)
        ]
        ws3.append(fila)
        excel_row = ws3.max_row
        ws3.cell(row=excel_row, column=7).number_format = '#,##0.00'
        ws3.cell(row=excel_row, column=8).number_format = '#,##0.00'
    
    ws3.column_dimensions["A"].width = 25
    ws3.column_dimensions["B"].width = 15
    ws3.column_dimensions["C"].width = 35
    ws3.column_dimensions["D"].width = 15
    ws3.column_dimensions["E"].width = 15
    ws3.column_dimensions["F"].width = 30
    ws3.column_dimensions["G"].width = 18
    ws3.column_dimensions["H"].width = 18
    
    wb.save(ruta_salida)

@router.post("/conciliar")
async def conciliar_archivos(
    file_dian: UploadFile = File(...),
    file_conta: UploadFile = File(...)
):
    """Endpoint para conciliar facturas DIAN con movimientos contables."""
    os.makedirs("temp", exist_ok=True)
    
    # Guardar archivos temporalmente usando los nombres del frontend
    ruta_facturas = f"temp/{uuid.uuid4()}_dian_{file_dian.filename}"
    ruta_movimientos = f"temp/{uuid.uuid4()}_conta_{file_conta.filename}"
    ruta_salida = f"temp/Conciliacion_{uuid.uuid4()}.xlsx"
    
    try:
        with open(ruta_facturas, "wb") as f:
            shutil.copyfileobj(file_dian.file, f)
        with open(ruta_movimientos, "wb") as f:
            shutil.copyfileobj(file_conta.file, f)
        
        # Procesar
        df_facturas = leer_facturas_dian(ruta_facturas)
        df_movimientos = leer_movimientos_contables(ruta_movimientos)
        
        if df_facturas.empty:
            raise HTTPException(status_code=400, detail="El archivo DIAN no contiene datos de facturas válidos.")
        if df_movimientos.empty:
            raise HTTPException(status_code=400, detail="El archivo de contabilidad no contiene datos válidos.")
        
        df_conciliacion = conciliar_por_tercero(df_facturas, df_movimientos)
        escribir_excel_conciliacion(ruta_salida, df_conciliacion, df_facturas, df_movimientos)
        
        return FileResponse(
            ruta_salida,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename="Conciliacion_DIAN_vs_Contabilidad.xlsx"
        )
    except HTTPException:
        raise  # Re-lanzar excepciones HTTP tal cual
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error inesperado procesando archivos: {str(e)}")
    finally:
        # Limpiar archivos de entrada
        for ruta in [ruta_facturas, ruta_movimientos]:
            if os.path.exists(ruta):
                os.remove(ruta)
        # Opcional: limpiar el archivo de salida después de un tiempo, 
        # pero FastAPI lo necesita para enviarlo, así que lo dejamos.