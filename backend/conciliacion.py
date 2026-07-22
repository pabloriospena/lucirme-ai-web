import os
import uuid
import shutil
import pandas as pd
from fastapi import APIRouter, File, UploadFile, HTTPException
from fastapi.responses import FileResponse
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

router = APIRouter()

# Colores
FILL_VERDE = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_AMARILLO = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_NARANJA = PatternFill(start_color="FCD5B4", end_color="FCD5B4", fill_type="solid")
FILL_ROJO = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_AZUL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
FILL_HEADER = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
FONT_HEADER = Font(color="FFFFFF", bold=True, size=11)

UMBRAL_COINCIDE = 1000
UMBRAL_DIFERENCIA_MENOR = 100000

def limpiar_numero(valor):
    if pd.isna(valor) or valor is None: return 0.0
    if isinstance(valor, (int, float)): return float(valor)
    try:
        return float(str(valor).replace(",", "").replace("$", "").strip())
    except (ValueError, TypeError):
        return 0.0

def limpiar_nit(valor):
    if pd.isna(valor) or valor is None: return ""
    return str(valor).replace(".", "").replace("-", "").replace(" ", "").strip()

def buscar_columna(df, palabras_clave):
    """Busca una columna que contenga alguna de las palabras clave, ignorando tildes y mayúsculas."""
    columnas_norm = {str(c).strip().lower().replace("á", "a").replace("é", "e").replace("í", "i").replace("ó", "o").replace("ú", "u"): c for c in df.columns}
    for key, original_name in columnas_norm.items():
        for palabra in palabras_clave:
            if palabra in key:
                return original_name
    return None

def leer_facturas_dian(ruta_archivo: str) -> pd.DataFrame:
    df = pd.read_excel(ruta_archivo, engine='openpyxl')
    
    # Buscar columnas dinámicamente
    col_tipo = buscar_columna(df, ["tipo de documento", "tipo documento"])
    col_nit = buscar_columna(df, ["nit emisor", "nit", "emisor"])
    col_nombre = buscar_columna(df, ["nombre emisor", "nombre", "emisor"])
    col_total = buscar_columna(df, ["total"])
    
    if not col_nit or not col_total:
        raise ValueError("No se encontraron las columnas 'NIT Emisor' o 'Total' en el archivo de facturas.")

    # Filtrar si existe la columna de tipo
    if col_tipo:
        tipos_validos = ["factura electronica", "nota de credito electronica", "documento equivalente pos", "factura electronica de contingencia"]
        df_filtrado = df[df[col_tipo].astype(str).str.lower().str.contains("|".join(tipos_validos), na=False)].copy()
    else:
        df_filtrado = df.copy()

    df_limpio = pd.DataFrame()
    df_limpio["NIT"] = df_filtrado[col_nit].apply(limpiar_nit)
    df_limpio["Nombre"] = df_filtrado[col_nombre] if col_nombre else "Desconocido"
    df_limpio["Total"] = df_filtrado[col_total].apply(limpiar_numero)
    
    return df_limpio

def leer_movimientos_contables(ruta_archivo: str) -> pd.DataFrame:
    df = pd.read_excel(ruta_archivo, engine='openpyxl')
    
    # Buscar columnas dinámicamente
    col_id = buscar_columna(df, ["identificacion", "nit", "cedula", "id"])
    col_nombre = buscar_columna(df, ["nombre tercero", "nombre", "tercero"])
    col_debito = buscar_columna(df, ["debito", "debe"])
    col_credito = buscar_columna(df, ["credito", "haber"])
    
    if not col_id:
        raise ValueError(f"No se encontró la columna de Identificación/NIT. Columnas encontradas: {list(df.columns)}")
    if not col_debito:
        raise ValueError("No se encontró la columna de Débito.")
    if not col_credito:
        raise ValueError("No se encontró la columna de Crédito.")

    df_limpio = pd.DataFrame()
    df_limpio["NIT"] = df[col_id].apply(limpiar_nit)
    df_limpio["Nombre"] = df[col_nombre] if col_nombre else "Desconocido"
    df_limpio["Debito"] = df[col_debito].apply(limpiar_numero)
    df_limpio["Credito"] = df[col_credito].apply(limpiar_numero)
    
    return df_limpio

def conciliar_por_tercero(df_facturas: pd.DataFrame, df_movimientos: pd.DataFrame) -> pd.DataFrame:
    # Agrupar facturas
    facturas_agg = df_facturas.groupby("NIT").agg(
        Nombre_Tercero=("Nombre", "first"),
        Total_Facturas=("Total", "sum")
    ).reset_index()
    
    # Agrupar movimientos
    movimientos_agg = df_movimientos.groupby("NIT").agg(
        Nombre_Tercero_Cont=("Nombre", "first"),
        Total_Debito=("Debito", "sum"),
        Total_Credito=("Credito", "sum")
    ).reset_index()
    movimientos_agg["Neto_Contable"] = movimientos_agg["Total_Debito"] - movimientos_agg["Total_Credito"]
    
    # Full outer join
    conciliacion = pd.merge(facturas_agg, movimientos_agg, on="NIT", how="outer", suffixes=("_Fac", "_Cont"))
    conciliacion["Nombre"] = conciliacion["Nombre_Tercero"].fillna(conciliacion["Nombre_Tercero_Cont"])
    
    # Llenar NaN con 0
    for col in ["Total_Facturas", "Total_Debito", "Total_Credito", "Neto_Contable"]:
        conciliacion[col] = conciliacion[col].fillna(0.0)
    
    conciliacion["Diferencia"] = (conciliacion["Total_Facturas"] - conciliacion["Neto_Contable"]).abs()
    
    def clasificar(row):
        if row["Total_Facturas"] == 0 and row["Neto_Contable"] != 0: return "Solo en Contabilidad"
        if row["Total_Facturas"] > 0 and row["Neto_Contable"] == 0: return "Solo en Facturación"
        if row["Diferencia"] < UMBRAL_COINCIDE: return "Coincide"
        if row["Diferencia"] < UMBRAL_DIFERENCIA_MENOR: return "Diferencia Menor"
        return "Diferencia"
    
    conciliacion["Estado"] = conciliacion.apply(clasificar, axis=1)
    
    orden_estado = {"Diferencia": 0, "Solo en Facturación": 1, "Solo en Contabilidad": 2, "Diferencia Menor": 3, "Coincide": 4}
    conciliacion["orden"] = conciliacion["Estado"].map(orden_estado)
    conciliacion = conciliacion.sort_values(["orden", "Diferencia"], ascending=[True, False])
    
    return conciliacion[["NIT", "Nombre", "Total_Facturas", "Total_Debito", "Total_Credito", "Neto_Contable", "Diferencia", "Estado"]].copy()

def escribir_excel_conciliacion(ruta_salida: str, df_conciliacion: pd.DataFrame, df_facturas: pd.DataFrame, df_movimientos: pd.DataFrame):
    wb = Workbook()
    
    # HOJA 1: CONCILIACIÓN
    ws1 = wb.active
    ws1.title = "Conciliación por Tercero"
    cols1 = ["Cédula / NIT", "Nombre de Tercero", "Total Facturas Electrónicas", "Total Débito Contable", "Total Crédito Contable", "Neto Contable (Deb - Cred)", "Estado de Conciliación", "Diferencia Mínima Estimada"]
    ws1.append(cols1)
    for c in range(1, len(cols1) + 1):
        cell = ws1.cell(row=1, column=c)
        cell.fill, cell.font, cell.alignment = FILL_HEADER, FONT_HEADER, Alignment(horizontal="center", vertical="center")
    
    estado_color = {"Coincide": FILL_VERDE, "Diferencia Menor": FILL_AMARILLO, "Diferencia": FILL_NARANJA, "Solo en Facturación": FILL_ROJO, "Solo en Contabilidad": FILL_AZUL}
    
    for _, row in df_conciliacion.iterrows():
        ws1.append([row["NIT"], row["Nombre"], row["Total_Facturas"], row["Total_Debito"], row["Total_Credito"], row["Neto_Contable"], row["Estado"], row["Diferencia"]])
        r = ws1.max_row
        color = estado_color.get(row["Estado"], FILL_VERDE)
        for c in range(1, 9):
            cell = ws1.cell(row=r, column=c)
            cell.fill = color
            if 3 <= c <= 8: cell.number_format = '#,##0.00'
            
    for col in ws1.columns:
        ws1.column_dimensions[col[0].column_letter].width = 25

    # HOJA 2: DETALLE FACTURAS
    ws2 = wb.create_sheet("Detalle Facturas (DIAN)")
    cols2 = ["Tipo de documento", "NIT Emisor", "Nombre Emisor", "Total", "Estado"]
    ws2.append(cols2)
    for c in range(1, len(cols2) + 1):
        cell = ws2.cell(row=1, column=c)
        cell.fill, cell.font = FILL_HEADER, FONT_HEADER
        
    col_tipo_orig = buscar_columna(df_facturas, ["tipo de documento"]) or "Tipo de documento"
    col_nit_orig = buscar_columna(df_facturas, ["nit emisor", "nit"]) or "NIT"
    col_nom_orig = buscar_columna(df_facturas, ["nombre emisor", "nombre"]) or "Nombre"
    col_tot_orig = buscar_columna(df_facturas, ["total"]) or "Total"
    col_est_orig = buscar_columna(df_facturas, ["estado"]) or "Estado"

    for _, row in df_facturas.iterrows():
        ws2.append([row.get(col_tipo_orig, ""), row.get(col_nit_orig, ""), row.get(col_nom_orig, ""), row.get(col_tot_orig, 0), row.get(col_est_orig, "")])
        ws2.cell(row=ws2.max_row, column=4).number_format = '#,##0.00'

    # HOJA 3: DETALLE MOVIMIENTOS
    ws3 = wb.create_sheet("Detalle Movimientos (Conta)")
    cols3 = ["Cuenta contable", "Identificación", "Nombre tercero", "Débito", "Crédito"]
    ws3.append(cols3)
    for c in range(1, len(cols3) + 1):
        cell = ws3.cell(row=1, column=c)
        cell.fill, cell.font = FILL_HEADER, FONT_HEADER
        
    col_cta_orig = buscar_columna(df_movimientos, ["cuenta contable", "cuenta"]) or "Cuenta contable"
    col_id_orig = buscar_columna(df_movimientos, ["identificacion", "nit", "cedula"]) or "NIT"
    col_nom_mov_orig = buscar_columna(df_movimientos, ["nombre tercero", "nombre"]) or "Nombre"
    col_deb_orig = buscar_columna(df_movimientos, ["debito", "debe"]) or "Debito"
    col_cre_orig = buscar_columna(df_movimientos, ["credito", "haber"]) or "Credito"

    for _, row in df_movimientos.iterrows():
        ws3.append([row.get(col_cta_orig, ""), row.get(col_id_orig, ""), row.get(col_nom_mov_orig, ""), row.get(col_deb_orig, 0), row.get(col_cre_orig, 0)])
        ws3.cell(row=ws3.max_row, column=4).number_format = '#,##0.00'
        ws3.cell(row=ws3.max_row, column=5).number_format = '#,##0.00'

    wb.save(ruta_salida)

@router.post("/conciliar")
async def conciliar_archivos(
    file_dian: UploadFile = File(...),
    file_conta: UploadFile = File(...)
):
    os.makedirs("temp", exist_ok=True)
    ruta_facturas = f"temp/{uuid.uuid4()}_dian_{file_dian.filename}"
    ruta_movimientos = f"temp/{uuid.uuid4()}_conta_{file_conta.filename}"
    ruta_salida = f"temp/Conciliacion_{uuid.uuid4()}.xlsx"
    
    try:
        with open(ruta_facturas, "wb") as f: shutil.copyfileobj(file_dian.file, f)
        with open(ruta_movimientos, "wb") as f: shutil.copyfileobj(file_conta.file, f)
        
        df_facturas = leer_facturas_dian(ruta_facturas)
        df_movimientos = leer_movimientos_contables(ruta_movimientos)
        
        if df_facturas.empty: raise HTTPException(status_code=400, detail="El archivo DIAN no contiene facturas válidas.")
        if df_movimientos.empty: raise HTTPException(status_code=400, detail="El archivo contable está vacío.")
        
        df_conciliacion = conciliar_por_tercero(df_facturas, df_movimientos)
        escribir_excel_conciliacion(ruta_salida, df_conciliacion, df_facturas, df_movimientos)
        
        return FileResponse(ruta_salida, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename="Reporte_Conciliacion.xlsx")
    except ValueError as ve:
        raise HTTPException(status_code=400, detail=str(ve))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error inesperado: {str(e)}")
    finally:
        for ruta in [ruta_facturas, ruta_movimientos]:
            if os.path.exists(ruta): os.remove(ruta)